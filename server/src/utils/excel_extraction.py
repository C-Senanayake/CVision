from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any
from datetime import datetime


def create_cv_excel(cv_data_list: List[Dict[str, Any]]) -> BytesIO:
    """
    Create an Excel file from CV data.
    
    Args:
        cv_data_list: List of CV documents from MongoDB
        
    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CV Export"
    
    # Define header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define headers
    headers = [
        "Candidate Name",
        "Email",
        "Phone",
        "Location",
        "LinkedIn",
        "GitHub",
        "Job Name",
        "Final Mark",
        "Selected for Interview",
        "University",
        "Degree",
        "Mail Status",
        "Interview Scheduled",
        "Interview Name",
        "Interview Location",
        "Interview Attendees",
        "Interview Start Datetime",
        "Interview End Datetime",
        "Created Date",
        "Updated Date"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row_num, cv in enumerate(cv_data_list, 2):
        resume_content = cv.get("resumeContent", {})
        personal_info = resume_content.get("personal_info", {})
        education = resume_content.get("education", [{}])[0] if resume_content.get("education") else {}
        interviewEvent = cv.get("interviewEvent", {})
        
        row_data = [
            cv.get("candidateName", ""),
            personal_info.get("email", ""),
            personal_info.get("phone", ""),
            personal_info.get("address", ""),
            personal_info.get("linkedin", ""),
            personal_info.get("github", ""),
            cv.get("jobName", ""),
            cv.get("finalMark", 0),
            "Yes" if cv.get("selectedForInterview", False) else "No",
            education.get("institution", ""),
            education.get("degree", ""),
            cv.get("mailStatus", ""),
            "Yes" if cv.get("interviewEvent", False) else "No",
            interviewEvent.get("interviewName", ""),
            interviewEvent.get("interviewLocation", ""),
            ", ".join(interviewEvent.get("interviewAttendees", [])),
            interviewEvent.get("interviewStartDatetime", ""),
            interviewEvent.get("interviewEndDatetime", ""),
            cv.get("createdAt", "").strftime("%Y-%m-%d %H:%M:%S") if cv.get("createdAt") else "",
            cv.get("updatedAt", "").strftime("%Y-%m-%d %H:%M:%S") if cv.get("updatedAt") else ""
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)  # Max width 60
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def format_cv_for_export(cv_document: Dict[str, Any]) -> Dict[str, Any]:
    """
    Format a single CV document for export.
    Can be used for data transformation if needed.
    
    Args:
        cv_document: Raw CV document from MongoDB
        
    Returns:
        Formatted CV data
    """
    return cv_document
